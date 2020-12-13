import os, requests, re, time, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from bs4 import BeautifulSoup
from progress.bar import IncrementalBar

# Set absolute file path for script and relative path for ticker file
script_path = os.path.dirname(__file__)
text_path = os.path.join(script_path, 'tickerList.txt')

# Get API key
key_file = open(script_path + "key.txt") # Open and read key file
key_string = key_file.read()
pattern = re.compile(r'[a-zA-Z0-9]+')
key = pattern.search(key_string).group() # store key value
key_file.close()

# Marketwatch URL
mark_watch = "https://www.marketwatch.com/investing/stock/"

# API values
api_url = "https://www.alphavantage.co/query?function=OVERVIEW&symbol="
key_url = "&apikey=" + key

# Spreadsheet formating values
title_font = Font(size = "14", bold = True, name = 'Calibri')
label_font = Font(name= 'Calibri', bold = True)
reg_font = Font(name = 'Calibri')
centered = Alignment(horizontal="center", vertical="center")

ticker_list = [] # List of stocks tickers to pull data on
stock_objs = [] # List of stock info as dictionaries
stock_prices = [] # List of stock prices taken from Marketwatch

# Get list of tickers
ticker_file = open(text_path)
string = ticker_file.read()
pattern = re.compile(r'[a-zA-Z]+')
find_result = pattern.findall(string)
for ticker in find_result:
    ticker_list.append(ticker)
ticker_file.close()


# Create function to populate stock_prices list
def get_prices(url):
    res = requests.get(url)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, 'html.parser')
    price = soup.select('body > div.container.container--body > div.region.region--intraday > div.column.column--aside > div.element.element--intraday > div.intraday__data > h3.intraday__price > bg-quote.value')
    try:
        stock_prices.append("$" + price[0].text)
    except:
        stock_prices.append("Not Found")

# Populate stock_prices list
price_bar = IncrementalBar('Getting Stock Prices', max = len(ticker_list)) # Progress bar for getting prices
for ticker in ticker_list:
    get_prices(mark_watch + ticker)
    price_bar.next()
    time.sleep(1)
price_bar.finish()

# Populate stock_objs
stock_bar = IncrementalBar('Getting Stock Overview Info', max = len(ticker_list)) # Progress bar for getting API info 
for ticker in ticker_list:
    response = requests.get(api_url + ticker + key_url)
    json_dict = json.loads(response.content)
    stock_objs.append(json_dict)
    stock_bar.next()
    time.sleep(1)
stock_bar.finish() # Progress bar shows finished

# Fill values
red = PatternFill("solid", fgColor="ffa6a6")
green = PatternFill("solid", fgColor="afd095")
yellow = PatternFill("solid", fgColor="ffffa6")
blue = PatternFill("solid", fgColor="b4c7dc")
orange = PatternFill("solid", fgColor="ffb66c")
purple = PatternFill("solid", fgColor="bf819e")

count = 1
f_count = 0
fill_list = [red, green, yellow, blue, orange, purple]

def populate_cells(a,b, ticker):
    # Creates and populates cells in the spreadsheet arguments are for the y range
    name = stock_objs[ticker]["Name"]
    eps = stock_objs[ticker]["EPS"]
    beta = stock_objs[ticker]["Beta"]
    pb = stock_objs[ticker]["PriceToBookRatio"]
    dps = stock_objs[ticker]["DividendPerShare"]
    pay_ratio = stock_objs[ticker]["PayoutRatio"]
    pe = stock_objs[ticker]["PERatio"]
    price = stock_prices[ticker]
    global f_count, count
    for x in range(count, count + 9):
        for y in range(a, b):
            if x == count: # Title cell
                if y == a:
                    cell = ws.cell(row=x, column=y, value= name) 
                    cell.font = title_font
                    cell.fill = fill_list[f_count]
            if x == count + 1: # Price cell
                if y == a:
                    cell = ws.cell(row=x, column=y, value= price)
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
            if x == count + 2: # Earnins Per Share
                if y == a:
                    cell = ws.cell(row=x, column=y, value= 'Earnings Per Share')
                    cell.font = reg_font
                    cell.fill = fill_list[f_count]
                if y == a + 1:
                    cell = ws.cell(row=x, column=y, value= eps)
                    cell.alignment = centered
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
            if x == count + 3: # Beta Ratio
                if y == a:
                    cell = ws.cell(row=x, column=y, value= 'Beta Ratio')
                    cell.font = reg_font
                    cell.fill = fill_list[f_count]
                if y == a + 1:
                    cell = ws.cell(row=x, column=y, value= beta)
                    cell.alignment = centered
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
            if x == count + 4: # Price to Book
                if y == a:
                    cell = ws.cell(row=x, column=y, value= 'Price to Book')
                    cell.font = reg_font
                    cell.fill = fill_list[f_count]
                if y == a + 1:
                    cell = ws.cell(row=x, column=y, value= pb)
                    cell.alignment = centered
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
            if x == count + 5: # Dividend Payment
                if y == a:
                    cell = ws.cell(row=x, column=y, value= 'Dividend Per Share')
                    cell.font = reg_font
                    cell.fill = fill_list[f_count]
                if y == a + 1:
                    cell = ws.cell(row=x, column=y, value= dps)
                    cell.alignment = centered
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
            if x == count + 6: # Dividend Yield %
                if y == a:
                    cell = ws.cell(row=x, column=y, value= 'Payout Ratio')
                    cell.font = reg_font
                    cell.fill = fill_list[f_count]
                if y == a + 1:
                    cell = ws.cell(row=x, column=y, value= pay_ratio)
                    cell.alignment = centered
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
            if x == count + 7: # Cash Ratio
                if y == a:
                    cell = ws.cell(row=x, column=y, value= 'Price to Earnings Ratio')
                    cell.font = reg_font
                    cell.fill = fill_list[f_count]
                if y == a + 1:
                    cell = ws.cell(row=x, column=y, value= pe)
                    cell.alignment = centered
                    cell.font = label_font
                    cell.fill = fill_list[f_count]
    count += 10
    f_count += 1
    if f_count > 4 :
        f_count = 0
    
# Create Workbook to be saved to "stockOutput.xlsx"
wb = Workbook()
ws = wb.active # Select default sheet
ws.title = "stockOutput" # Rename sheet title

pop_cell_bar = IncrementalBar('Creating Workbook', max = len(stock_objs)) # Populing workbook cells progress
# Populate first columns of data
for i in range(0, len(stock_objs) // 2): 
    populate_cells(1,3, i)
    pop_cell_bar.next()
    time.sleep(1)

count = 1
# Populate second columns of data
for i in range(len(stock_objs) // 2, len(stock_objs)):
    populate_cells(4,6, i)
    pop_cell_bar.next()
    time.sleep(1)
pop_cell_bar.finish() # Populating Cells finished.

ws.column_dimensions['a'].width = 50 # Set Column A width
ws.column_dimensions['b'].width = 10 # Set Column B width
ws.column_dimensions['d'].width = 50 # Set Column D width
ws.column_dimensions['e'].width = 10 # Set Column E width

wb.save("stockOutput.xlsx") # Save out file. NOTE: Overwrites entire file everytime script runs.
